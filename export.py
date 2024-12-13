
import openpyxl
from workers import ROR
import pythoncom
import win32com.client
from base import get_form_data


def insert_data_to_excel(file_name, data_dict, data_netipovye_raboty, user_id):
    try:
        workbook = openpyxl.load_workbook(file_name, keep_vba=True)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    for sheet_name in ('1', '3', '4 гор', '5-8 гор', '9-10', '11'):
        worksheet = workbook[sheet_name] # Получение листа

        # Соответствие ключей словаря ячейкам
        cell_mapping = {}

        if sheet_name == '1':
            worksheet["A11"] = data_dict.get('адрес')
            mark = ()
            cell_mapping = {
                "РОР": "B16",
                "рп": "B18",
                "ИСК": "B20",

            }
        

        elif sheet_name == '3':
            worksheet["C30"] = data_dict.get("рп")

            # тип объекта:
            if data_dict.get("тип_объекта") == "Встроен./встроен.-пристроен.":
                tipe_of_object = "E7"
            elif data_dict.get("тип_объекта") == "Торг. Центр":
                tipe_of_object = "E8"
            elif data_dict.get("тип_объекта") == "Цоколь/подвал. Этаж":
                tipe_of_object = "E9"
            else:
                tipe_of_object = "E10"


            if data_dict.get("использование_подвала") == "Да":
                using_base = "H12"
            else:
                using_base = "J12"

            if data_dict.get("соответствие_планировки") == "Да":
                compliance = "E15"
            else:
                compliance = "G15"


            mark = ("E7", "E8", "E9", "E10", "H12", "J12", "E15", "G15")

            # Очищаем ранее заполненные данные
            for cell in mark:
                worksheet[cell] = ''

            cell_mapping = {
                "площадь_помещенияь": "C4",
                "этаж": "B5",
                "этажность": "E5",
                "тип_объекта": tipe_of_object,
                "использование_подвала": using_base,
                "комментарий_6": "B13",
                "соответствие_планировки": compliance,
                "комментарий_7": "B16",
                "фундамент": "B20",
                "полы": "B21",
                "нагрузка": "E21",
                "стены": "B22",
                "тип_потолка": "C23",
                "материал_потолка": "H23",
                "тип_пола": "C24",
                "материал_пола": "H24",
                "кровля": "B25",
                "нагрузка_кровли": "E25",
                "конструктивная_схема": "C27",
                "дефекты": "C28",

                }

        elif sheet_name == '4 гор':
            worksheet['B34'] = data_dict.get('рп')

        elif sheet_name == '5-8 гор':
            worksheet['B37'] = data_dict.get('рп')
            worksheet['B70'] = data_dict.get('рп')
            worksheet['B108'] = data_dict.get('рп')

        elif sheet_name == '9-10':
            worksheet['B22'] = data_dict.get('рп')
            worksheet['B24'] = data_dict.get('член_ком1')
            worksheet['B55'] = data_dict.get('рп')

        elif sheet_name == '11':
            worksheet['A12'] = data_dict.get('')
            worksheet['B43'] = data_dict.get('рп')
            # worksheet['B44'] = data_dict.get('председатель')
            mark = (
                    'B6', 'B10',
                    )
 
            # Очищаем ранее заполненные данные
            for cell in mark:
                worksheet[cell] = ''

            for row in range(21, 34):
                for col in (1, 2, 3, 4, 5, 9):
                    worksheet.cell(row=row, column=col).value = ''   

            # ВЫВОД: Использование помещений/здания в качестве магазина ТС «Монетка»:
            if data_dict.get("возможность") == "Возможно":
                possibly = "B6"
            else:
                possibly = "B10"

            cell_mapping = {
                "возможность" : possibly,
                "причина_невозможности": "A12",
                "работы_не_требующие": "A15",
                "требования" : "A37",
                "срок_строительства" : "D41"
            }


            # Заполняем таблицу
            if get_form_data(user_id)["нетиповые_работы"] == 'Да':
                for i, item in enumerate(data_netipovye_raboty.get("нетиповые_работы")):
                    worksheet['A' + str(21 + i)] = i + 1
                    worksheet['B' + str(21 + i)] = item.get("тип_работ")
                    if item.get("срок") == "до АПП":
                        worksheet['C' + str(21 + i)] = "X"
                    elif item.get("срок") == "до ВПК":
                        worksheet['D' + str(21 + i)] = "X"
                    else:
                        worksheet['E' + str(21 + i)] = "X"
                    worksheet['I' + str(21 + i)] = item.get("ответственный")

        # Запись данных в ячейки
        for key, value in data_dict.items():
            if key in cell_mapping.keys():
                cell = cell_mapping[key]
                if value == 'Пропущено':
                    pass
                else:
                    if cell in mark:
                        worksheet[cell] = 'X'
                    else:
                        worksheet[cell] = value

    workbook.save(file_name)  # Сохранение изменений
    print(f"Данные успешно вставлены в файл '{file_name}'.")


def xlsm_to_pdf(xlsm_path):
    try:
        pythoncom.CoInitialize()
        Excel = win32com.client.Dispatch("Excel.Application")
        Excel.Visible = 0
        wb = Excel.Workbooks.Open(xlsm_path)
        file_pdf = xlsm_path.split(".xlsm")[0] + ".pdf"
        wb.ExportAsFixedFormat(0, file_pdf.replace('/', '\\'))
        wb.Close()
    except Exception as e:
        print(e)
    finally:
        Excel.Quit()
